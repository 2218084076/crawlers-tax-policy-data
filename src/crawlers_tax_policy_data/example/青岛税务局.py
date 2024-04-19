import sys
import time
from datetime import datetime

import requests
import undetected_chromedriver as uc
from lxml import etree
from openpyxl import Workbook


def get_cookie():
    driver = uc.Chrome()

    url = "http://qingdao.chinatax.gov.cn/ssfg2019/zxwj/202404/t20240410_86667.html"
    driver.get(url)
    time.sleep(3)
    cookies = driver.get_cookies()
    new_cookies = ''
    for cookie in cookies:
        name = cookie['name']
        value = cookie['value']

        items = f'{name}={value};'
        new_cookies += items

    print(new_cookies)
    # 设置请求头中的 Cookie
    headers['Cookie'] = new_cookies
    driver.close()


def get_list(page):
    global contxx

    if page == 0:
        url = 'http://qingdao.chinatax.gov.cn/ssfg2019/zxwj/index.html'
    else:
        url = f'http://qingdao.chinatax.gov.cn/ssfg2019/zxwj/index_{page}.html'

    resp = requests.get(url, headers=headers, proxies=proxy)
    resp.encoding = 'utf-8'
    resp = resp.text

    tree = etree.HTML(resp)

    lis = tree.xpath('//ul[@class="pc_normal_content_ul"]/li')

    for li in lis:

        title = li.xpath('./a/text()')
        shijian = li.xpath('./span/text()')

        title = "".join(title).strip()
        shijian = "".join(shijian).strip()

        new_shijian = "".join(shijian)[:10]
        yue = "".join(shijian)[:7].replace('-', '')

        reference_date = datetime.strptime(new_shijian, '%Y-%m-%d')

        print(end_time)
        print(reference_date)
        if end_time > reference_date:
            # 结束程序
            sys.exit()

        href = li.xpath('./a/@href')
        href = "".join(href).split('./')[-1].strip()
        x_url = f'http://qingdao.chinatax.gov.cn/ssfg2019/zxwj/{href}'

        print(x_url, title, shijian)
        while True:
            try:
                resp = requests.get(x_url, headers=headers, proxies=proxy)
                resp.encoding = 'utf-8'
                resp = resp.text

                tree = etree.HTML(resp)
                wenshu = tree.xpath('//div[@class="pc_nrxq1_tittle"]/div[3]/text()')
                wenshu = "".join(wenshu).strip()

                texts = tree.xpath('//div[@class="pc_textcontent"]//text()')
                texts = "".join(texts).strip()
                print(texts)

                youxiao = tree.xpath('//div[@class="pc_textcontent"]//p[@align="right"]//text()')
                youxiao = "".join(youxiao).strip()

                if '全文有效' in youxiao:
                    youxiaoyn = '全文有效'
                else:
                    youxiaoyn = '-'

                if texts == '':
                    get_cookie()
                    print(texts, 'cookie失效,重新获取中')

                texts_html = tree.xpath('//div[@class="pc_textcontent"]')[0]
                texts_html = etree.tostring(texts_html, encoding='unicode')

                print(texts)
                ass = tree.xpath('//div[@class="pc_textcontent"]//a')
                ass_list = []
                for a in ass:
                    href = a.xpath('./@href')
                    p = a.xpath('./text()')
                    href = "".join(href).strip().split('./')[-1]
                    p = "".join(p).strip()

                    fj_url = f'http://qingdao.chinatax.gov.cn/ssfg2019/zxwj/{yue}/{href}'
                    ass_list.append(fj_url)
                    ass_list.append(p)
                    print(fj_url, p)

                jiedus = tree.xpath('//div[@id="xgzc"]/a')
                jiedus_list = []
                for jiedu in jiedus:
                    href = jiedu.xpath('./@href')
                    p = jiedu.xpath('./text()')
                    href = "".join(href).strip().split('./')[-1]
                    p = "".join(p).strip()

                    jd_url = f'http://qingdao.chinatax.gov.cn/ssfg2019/{href}'
                    jiedus_list.append(jd_url)
                    jiedus_list.append(p)
                    print(jd_url, p)
                if ass_list == []:
                    ass_list = '-'
                if jiedus_list == []:
                    jiedus_list = '-'

                break

            except:
                print('重试')
                time.sleep(3)
        ws[f'a{contxx}'] = x_url
        ws[f'b{contxx}'] = title
        ws[f'c{contxx}'] = wenshu
        ws[f'd{contxx}'] = youxiaoyn
        ws[f'e{contxx}'] = shijian
        ws[f'f{contxx}'] = '-'
        ws[f'g{contxx}'] = texts
        ws[f'h{contxx}'] = texts_html
        ws[f'i{contxx}'] = str(ass_list)
        ws[f'j{contxx}'] = str(jiedus_list)
        contxx += 1
        # 保存Excel文件
        wb.save(f'青岛税务{t}.xlsx')

        time.sleep(3)


if __name__ == '__main__':

    # 您想要检查的日期
    end_time = '2024-03-27'
    end_time = datetime.strptime(end_time, '%Y-%m-%d')

    t = int(time.time())

    wb = Workbook()
    ws = wb.active

    # ws[f'a{1}'] = 'url'
    # ws[f'B{1}'] = '标题'
    # ws[f'c{1}'] = '文号'
    # ws[f'd{1}'] = '发布日期'
    # ws[f'e{1}'] = '正文'
    # ws[f'f{1}'] = 'html代码'
    # ws[f'g{1}'] = '附件'
    contxx = 1

    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Cache-Control": "max-age=0",
        # "Cookie": "zh_choose=s; zh_choose=s; gizqLgxJ4gkhS=60nV56c74xcYDH0VOfQzKNg1LPEw50HBqZVibfYs3mNFr4sfkJKExD1q5zRehL6G7KkU0DU8sCROjI8Y73auGkga; _trs_uv=lv22sacn_343_hsmb; acw_tc=7ceef50a17133329879078821e5e0cff15a0b8ab4e827dec74af9e7fab; _trs_ua_s_1=lv3e80x2_343_t2v; _yfxkpy_ssid_10003717=%7B%22_yfxkpy_firsttime%22%3A%221713253303095%22%2C%22_yfxkpy_lasttime%22%3A%221713332986047%22%2C%22_yfxkpy_visittime%22%3A%221713332986047%22%2C%22_yfxkpy_domidgroup%22%3A%221713332986047%22%2C%22_yfxkpy_domallsize%22%3A%22100%22%2C%22_yfxkpy_cookie%22%3A%2220240416154143096361648508662227%22%2C%22_yfxkpy_returncount%22%3A%221%22%7D; gizqLgxJ4gkhT=0dGrQiQEcsk.lM2oabAQvHjX7nPfTl0wJVG7rMavFIEIc6xwEOKGtPjs93gxughvRs1DGC42GVLGz0DXWwlandXcCAguvRA.Gj1q5OlRb933SwtZmHj6lSAlQA.f.EwCSU5koTPTVW4szoBgcbn4BvRyyJQ4X_C8yLyS_Z3ymu6i.VwsPl6r_qqPxLsrrCAythyrIayhirq6MnwmVhBVr6pIyby0hQ78IsrBlJUY26Dy4rLXgGHi4Kq.1CaFygVoIi4RxTPEdQPfHSxbu4l7tQYHo3lfqy314qvggYOHldTtF7VJdiiFsaWLEDi4auBMq1Yfoa6e7c2nGI_ZbJEySDuJO1irHPyy8zvmxZ2HcJ7RmaJO1v5rRWhg3QNIliWrC3qwKR1iT_KfyDqbt81VI29JtwPhkLqWYK.UAtrxwtpW",
        "Host": "qingdao.chinatax.gov.cn",
        "Proxy-Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    }
    proxy = {'https': 'http://127.0.0.1:7890', 'http': 'http://127.0.0.1:7890'}

    get_cookie()

    for page in range(0, 2):
        get_list(page)
